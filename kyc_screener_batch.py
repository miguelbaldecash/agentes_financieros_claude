"""
KYC SCREENER — BATCH ANALYSIS
==============================
Parametros de entrada:
  - Solicitudes de sueldo de trabajo (tipo_renta_id = 1, BOLETA DE PAGO)
  - Sin apoyo familiar (recibe_boleta_rh_pariente IS NULL OR = 0)
  - CON cupon de activador (cupon_descuento_id IS NOT NULL) — FILTRO DURO
  - Aprobadas y firmadas (firmado = 1)
  - Firmadas en los ultimos 6 meses
  - Personas con 2+ solicitudes creadas en los ultimos 12 meses
  - Con al menos 1 documento OCR procesado

Analisis:
  - 14 reglas de deteccion (R1-R14)
  - R1-R8: Cross-solicitud (OCR profundo)
  - R9-R14: Individual
  - Clasificacion BAJO/MEDIO/ALTO buscando ~33% cada grupo
  - FPD 30+ por grupo

Output:
  - KYC_BATCH_120_CASOS.xlsx (4 hojas)
"""

import subprocess, os, re, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter, defaultdict
from statistics import quantiles

# ============================================================
# CONFIG
# ============================================================
MYSQL = r"C:\Program Files\MySQL\MySQL Workbench 8.0 CE\mysql.exe"
MA = [MYSQL, "-h", "databalde-cashsys.cloa3om0c8wi.us-west-2.rds.amazonaws.com",
      "-u", "readonly_miguel_zavala", "-puruUmbwslj4CNz", "-D", "databalde-cashsys",
      "--default-character-set=utf8", "-N", "-B"]

PSQL = r"C:\Program Files\PostgreSQL\17\bin\psql.exe"
PSQL_CONN = "host=airtable-sync.cluster-ckkmmadckjbw.us-east-1.rds.amazonaws.com port=5432 dbname=postgres user=claude_lectura sslmode=require"

OUTPUT = r"C:\Users\USER\Documents\Agente de Finanzas\KYC_BATCH_120_CASOS.xlsx"

# DNIs de representantes BaldeCash que aparecen en contratos — excluir de R4
DNIS_BALDECASH = {'10314652', '14402096', '14511932'}  # Montenegro, rep legal 1, rep legal 2

# ============================================================
# HELPERS
# ============================================================
def mq(sql):
    r = subprocess.run(MA + ["-e", sql], capture_output=True, timeout=300)
    return [l.split('\t') for l in r.stdout.decode('utf-8', errors='replace').strip().split('\n') if l]

def sf(v):
    try: return float(v) if v and v not in ('NULL', '') else None
    except: return None

def si(v):
    try: return int(v) if v and v not in ('NULL', '') else None
    except: return None

def get_pg_token():
    cmd = r'''powershell.exe -Command "& 'C:\Program Files\Amazon\AWSCLIV2\aws.exe' rds generate-db-auth-token --hostname airtable-sync.cluster-ckkmmadckjbw.us-east-1.rds.amazonaws.com --port 5432 --region us-east-1 --username claude_lectura"'''
    return subprocess.run(cmd, capture_output=True, text=True, shell=True, timeout=30).stdout.strip().replace('\r', '')

def pg_query(sql, token):
    env = os.environ.copy()
    env['PGPASSWORD'] = token
    res = subprocess.run([PSQL, PSQL_CONN, "-t", "-A", "-F", "\t", "-c", sql],
                        capture_output=True, text=True, env=env, timeout=60)
    return res.stdout.strip()

# ============================================================
# STEP 1: PERSONA IDS
# Filtro: boleta, sin apoyo familiar, sin cupon, firmada 6m, 2+ sols 12m, con OCR
# ============================================================
print("1/6 Personas...")
QUERY_PERSONAS = """
SELECT DISTINCT sol2.id_persona
FROM solicitud s2
JOIN solicitante sol2 ON s2.id_solicitante = sol2.id
WHERE s2.firmado = 1
  AND s2.tipo_renta_id = 1
  AND s2.airtable_fecha_firma >= DATE_SUB(NOW(), INTERVAL 6 MONTH)
  AND s2.cupon_descuento_id IS NOT NULL
  AND (s2.recibe_boleta_rh_pariente IS NULL OR s2.recibe_boleta_rh_pariente = 0)
  AND sol2.id_persona IN (
    SELECT sol3.id_persona FROM solicitud s3
    JOIN solicitante sol3 ON s3.id_solicitante = sol3.id
    WHERE s3.created_at >= DATE_SUB(NOW(), INTERVAL 12 MONTH)
    GROUP BY sol3.id_persona HAVING COUNT(DISTINCT s3.id) >= 2
  )
  /* Sin filtro de OCR — se usa OCR local para los faltantes */
"""
pid_rows = mq(QUERY_PERSONAS)
pids = [r[0] for r in pid_rows]
print(f"  {len(pids)} personas")

# ============================================================
# STEP 2: ALL SOLICITUDES
# ============================================================
print("2/6 Solicitudes...")
all_rows = []
for i in range(0, len(pids), 100):
    batch = pids[i:i+100]
    ids_str = ','.join(batch)
    q = f"""SELECT p.documento, CONCAT(p.nombres,' ',p.ape_pat,' ',p.ape_mat), sol.id_persona,
      s.id, s.status, s.edad, s.sueldo_bruto, s.sueldo_neto, s.pension_mensual,
      s.ratio_cuota_ingreso, s.score_final, s.ocupacion, s.tipo_renta_id,
      s.firmado, s.entregado_txt, s.caso_especial, s.departamento,
      s.created_at, s.airtable_fecha_firma, s.equifax_request_id,
      s.keynua_error, s.keynua_dni_mismatch, s.arbol_decision_aplicado, s.cupon_descuento_id
    FROM solicitud s JOIN solicitante sol ON s.id_solicitante=sol.id JOIN persona p ON sol.id_persona=p.id
    WHERE sol.id_persona IN ({ids_str}) ORDER BY sol.id_persona, s.created_at"""
    all_rows.extend(mq(q))
print(f"  {len(all_rows)} solicitudes totales")

personas = defaultdict(list)
for r in all_rows:
    personas[r[2]].append(r)

# Target: ultima firmada con boleta, sin cupon, sin apoyo familiar, en 6m
target_sols = {}
for pid, sols in personas.items():
    firmadas = [s for s in sols if s[13] == '1' and s[18] and s[18] >= '2026-02-01'
                and s[12] == '1' and s[23] not in ('NULL', '')]
    if firmadas:
        target_sols[pid] = firmadas[-1]
print(f"  {len(target_sols)} targets")

# ============================================================
# STEP 3: OCR
# ============================================================
print("3/6 OCR...")
all_sol_ids = set(s[3] for sols in personas.values() for s in sols)
ocr_map = defaultdict(list)
batch_ids = list(all_sol_ids)
for i in range(0, len(batch_ids), 150):
    chunk = batch_ids[i:i+150]
    ids_str = ','.join(chunk)
    q_ocr = f"""SELECT d.id_solicitud,
      JSON_UNQUOTE(JSON_EXTRACT(d.ocr_producto,'$.nombre_cliente')),
      JSON_UNQUOTE(JSON_EXTRACT(d.ocr_producto,'$.nombre_emisor')),
      JSON_UNQUOTE(JSON_EXTRACT(d.ocr_producto,'$.monto_total')),
      JSON_UNQUOTE(JSON_EXTRACT(d.ocr_producto,'$.tipo_documento')),
      JSON_UNQUOTE(JSON_EXTRACT(d.ocr_producto,'$.ruc_emisor')),
      JSON_UNQUOTE(JSON_EXTRACT(d.ocr_producto,'$.datos.dni')),
      JSON_UNQUOTE(JSON_EXTRACT(d.ocr_producto,'$.datos.score')),
      JSON_UNQUOTE(JSON_EXTRACT(d.ocr_producto,'$.datos.rci'))
    FROM documentacion d WHERE d.deleted_at IS NULL AND d.ocr_producto IS NOT NULL AND d.id_solicitud IN ({ids_str})"""
    for ro in mq(q_ocr):
        if any(v not in ('NULL', '', None) for v in ro[1:]):
            ocr_map[ro[0]].append({
                'nombre': ro[1] if ro[1] not in ('NULL', '') else None,
                'empleador': ro[2] if ro[2] not in ('NULL', '') else None,
                'monto': sf(ro[3]),
                'tipo': ro[4] if ro[4] not in ('NULL', '') else None,
                'ruc': ro[5] if ro[5] not in ('NULL', '') else None,
                'ocr_dni': ro[6] if ro[6] not in ('NULL', '') else None,
                'ocr_score': ro[7] if ro[7] not in ('NULL', '') else None,
                'ocr_rci': ro[8] if ro[8] not in ('NULL', '') else None,
            })
print(f"  {sum(len(v) for v in ocr_map.values())} docs OCR (DB)")

# Load local OCR results for missing personas
LOCAL_OCR = r"C:\Users\USER\Documents\Agente de Finanzas\ocr_faltantes_resultado.json"
if os.path.exists(LOCAL_OCR):
    with open(LOCAL_OCR, 'r', encoding='utf-8') as f:
        local_ocr = json.load(f)
    local_added = 0
    for sol_id, docs_list in local_ocr.items():
        if sol_id not in ocr_map or len(ocr_map[sol_id]) == 0:
            for d in docs_list:
                entry = {
                    'nombre': None, 'empleador': None, 'monto': None,
                    'tipo': d.get('tipo_detectado'), 'ruc': d.get('ruc'),
                    'ocr_dni': None, 'ocr_score': None, 'ocr_rci': None,
                }
                if d.get('nombre_line'): entry['nombre'] = d['nombre_line']
                if d.get('empleador_line'): entry['empleador'] = d['empleador_line']
                if d.get('montos') and len(d['montos']) > 0: entry['monto'] = d['montos'][0]
                if d.get('dnis_found') and len(d['dnis_found']) > 0: entry['ocr_dni'] = d['dnis_found'][0]
                if d.get('score'): entry['ocr_score'] = str(d['score'])
                if d.get('rci'): entry['ocr_rci'] = str(d['rci'])
                ocr_map[sol_id].append(entry)
                local_added += 1
    print(f"  + {local_added} docs OCR local (faltantes)")
else:
    print("  [Sin OCR local]")

# ============================================================
# STEP 4: FPD
# ============================================================
print("4/6 FPD...")
token = get_pg_token()
fpd_map = {}
target_ids = [target_sols[pid][3] for pid in target_sols]
for i in range(0, len(target_ids), 50):
    batch = target_ids[i:i+50]
    sql = f"SET client_encoding TO 'UTF8'; SELECT loan_id, MAX(CASE WHEN month_on_book=2 THEN eom_days_delinquent END), MAX(CASE WHEN month_on_book=3 THEN eom_days_delinquent END) FROM loantape_periodo WHERE tipo_precio='compra' AND effective_to IS NULL AND loan_id IN ({','.join(batch)}) GROUP BY loan_id;"
    for line in pg_query(sql, token).split('\n'):
        if line and '\t' in line:
            p = line.split('\t')
            fpd_map[p[0].strip()] = {'dpd2': si(p[1].strip()) if len(p) > 1 else None}
print(f"  {len(fpd_map)} con FPD")

# ============================================================
# STEP 5: SCORING (14 reglas)
# ============================================================
print("5/6 Scoring...")

def score14(pid, sols, ocr_map):
    anomalies = []; pts = 0
    nombre_sol = sols[0][1].upper()
    dni_sol = sols[0][0]
    ape_parts = [p for p in nombre_sol.split() if len(p) > 3]
    target = sols[-1]

    sueldos = []; all_empleadores = set(); all_nombres_ocr = set(); all_rucs = set()
    montos_boleta = defaultdict(list); dnis_ocr = set(); scores_ocr = []; rcis_ocr = []

    for s in sols:
        sn = sf(s[7])
        if sn and sn > 0: sueldos.append(sn)
        for o in ocr_map.get(s[3], []):
            if o['empleador']: all_empleadores.add(o['empleador'].upper().strip())
            if o['nombre']: all_nombres_ocr.add(o['nombre'].upper().strip())
            if o['ruc']: all_rucs.add(o['ruc'])
            if o['monto'] and o['tipo'] == 'boleta_venta': montos_boleta[s[3]].append(o['monto'])
            if o['ocr_dni']: dnis_ocr.add(o['ocr_dni'])
            if o['ocr_score']:
                try: scores_ocr.append(int(o['ocr_score']))
                except: pass
            if o['ocr_rci']:
                try:
                    rv = float(o['ocr_rci'].replace('%', ''))
                    if rv > 1: rv /= 100
                    rcis_ocr.append(rv)
                except: pass

    # R1: Sueldo varia
    if len(sueldos) >= 2:
        mx, mn = max(sueldos), min(sueldos)
        if mn > 0 and mx/mn > 2.0: pts += 3; anomalies.append(f"R1: Sueldo {mn:.0f}-{mx:.0f} ({mx/mn:.1f}x)")
        elif mn > 0 and mx/mn > 1.5: pts += 2; anomalies.append(f"R1: Sueldo {mn:.0f}-{mx:.0f} ({mx/mn:.1f}x)")
        elif mn > 0 and mx/mn > 1.3: pts += 1; anomalies.append(f"R1: Sueldo {mn:.0f}-{mx:.0f}")

    # R2: Empleadores distintos
    unique_emps = []
    for e in all_empleadores:
        if not any(e[:8] == u[:8] or u in e or e in u for u in unique_emps): unique_emps.append(e)
    if len(unique_emps) >= 3: pts += 3; anomalies.append(f"R2: {len(unique_emps)} empleadores")
    elif len(unique_emps) == 2: pts += 2; anomalies.append(f"R2: 2 empleadores")

    # R3: Nombres OCR != solicitante
    nombres_no_match = set()
    for n in all_nombres_ocr:
        if not any(p in n for p in ape_parts): nombres_no_match.add(n)
    if len(nombres_no_match) >= 2: pts += 4; anomalies.append(f"R3: Docs de {len(nombres_no_match)} personas")
    elif len(nombres_no_match) == 1: pts += 2; anomalies.append(f"R3: Doc a nombre de otro")

    # R4: DNI en doc != solicitante (excluir DNIs de representantes BaldeCash)
    dnis_mismatch = [d for d in dnis_ocr if d != dni_sol and len(d) >= 7 and d not in DNIS_BALDECASH]
    if dnis_mismatch: pts += 3; anomalies.append(f"R4: DNI doc != solicitante")

    # R5: Boletas vs sueldo declarado
    sueldo_decl = sf(target[7]) or 0
    target_boletas = montos_boleta.get(target[3], [])
    if target_boletas and sueldo_decl > 0 and len(target_boletas) >= 3:
        avg_b = sum(target_boletas) / len(target_boletas)
        monthly = avg_b * 4
        if abs(monthly - sueldo_decl) / sueldo_decl > 0.5:
            pts += 2; anomalies.append(f"R5: Boletas vs sueldo")

    # R6: RCI en doc > 40%
    if rcis_ocr and max(rcis_ocr) > 0.40: pts += 2; anomalies.append(f"R6: RCI {max(rcis_ocr):.0%}")

    # R7: Equifax score OCR bajo
    if scores_ocr and min(scores_ocr) < 400: pts += 2; anomalies.append(f"R7: EQF OCR {min(scores_ocr)}")
    elif scores_ocr and min(scores_ocr) < 500: pts += 1; anomalies.append(f"R7: EQF OCR {min(scores_ocr)}")

    # R8: Multiples RUCs
    if len(all_rucs) >= 3: pts += 2; anomalies.append(f"R8: {len(all_rucs)} RUCs")
    elif len(all_rucs) == 2: pts += 1; anomalies.append("R8: 2 RUCs")

    # R9: Ratio rechazo
    n_total = len(sols); n_rej = sum(1 for s in sols if s[4] == 'rechazado')
    if n_total >= 5 and n_rej/n_total > 0.7: pts += 2; anomalies.append(f"R9: Rechazo {n_rej}/{n_total}")
    elif n_total >= 8: pts += 1; anomalies.append(f"R9: {n_total} sols")

    # R10: Edad
    edad_t = si(target[5]) or 0
    if edad_t > 0 and edad_t < 19: pts += 2; anomalies.append(f"R10: Edad {edad_t}")
    elif edad_t > 0 and edad_t < 20: pts += 1; anomalies.append(f"R10: Edad {edad_t}")

    # R11: Score
    score_f = si(target[10])
    if score_f and score_f < 250: pts += 2; anomalies.append(f"R11: Score {score_f}")
    elif score_f and score_f < 350: pts += 1; anomalies.append(f"R11: Score {score_f}")
    elif score_f is None: pts += 1; anomalies.append("R11: Sin score")

    # R12: Sueldo
    if sueldo_decl < 800 and sueldo_decl > 0: pts += 1; anomalies.append(f"R12: Sueldo S/{sueldo_decl:.0f}")
    elif sueldo_decl == 0: pts += 2; anomalies.append("R12: Sueldo=0")

    # R13: Controles
    if target[19] in ('NULL', None, '', '0'): pts += 1; anomalies.append("R13: Sin Equifax")
    if target[20] == '1': pts += 2; anomalies.append("R13: Keynua error")
    if target[21] == '1': pts += 3; anomalies.append("R13: Keynua DNI mismatch")
    if target[22] in ('NULL', None, ''): pts += 1; anomalies.append("R13: Sin arbol")

    # R14: Edad varia
    edades = [si(s[5]) for s in sols if si(s[5]) and si(s[5]) > 0]
    if edades and max(edades) - min(edades) > 3: pts += 1; anomalies.append(f"R14: Edad {min(edades)}-{max(edades)}")

    n_docs_ocr = sum(len(ocr_map.get(s[3], [])) for s in sols)
    n_boletas = sum(1 for s in sols for o in ocr_map.get(s[3], []) if o['tipo'] == 'boleta_venta')

    return pts, anomalies, {
        'n_docs_ocr': n_docs_ocr, 'n_boletas': n_boletas,
        'n_empleadores': len(unique_emps), 'n_nombres_mismatch': len(nombres_no_match),
        'n_dnis_mismatch': len(dnis_mismatch), 'equifax_score_ocr': min(scores_ocr) if scores_ocr else None,
    }

results = []
for pid in target_sols:
    target = target_sols[pid]; sols = personas[pid]
    pts, anomalies, stats = score14(pid, sols, ocr_map)
    dpd = fpd_map.get(target[3], {})
    dpd2 = dpd.get('dpd2')
    fpd30 = 1 if dpd2 is not None and dpd2 > 30 else (0 if dpd2 is not None else None)
    results.append({
        'dni': target[0], 'nombre': target[1], 'pid': pid, 'sol_id': target[3],
        'edad': si(target[5]) or 0, 'sueldo': sf(target[7]) or 0,
        'rci': sf(target[9]), 'score': si(target[10]),
        'ocupacion': target[11] if target[11] not in ('NULL', '-', '') else '-',
        'depto': target[16] if target[16] not in ('NULL', '-', '') else '-',
        'n_sols': len(sols), 'n_firmadas': sum(1 for s in sols if s[13] == '1'),
        'n_rechazadas': sum(1 for s in sols if s[4] == 'rechazado'),
        'dpd2': dpd2, 'fpd30': fpd30, 'has_loan': target[3] in fpd_map,
        'pts': pts, 'anomalies': '; '.join(anomalies) if anomalies else 'Ninguna',
        'n_anomalies': len(anomalies), **stats,
    })

# ============================================================
# STEP 6: CLASSIFY + EXCEL
# ============================================================
print("6/6 Clasificando y Excel...")
n = len(results)
# Clasificacion real basada en puntos de anomalia (sin forzar distribucion)
t1, t2 = 3, 5
for r in results:
    if r['pts'] <= t1: r['risk'] = 'BAJO'
    elif r['pts'] <= t2: r['risk'] = 'MEDIO'
    else: r['risk'] = 'ALTO'

bajo = sum(1 for r in results if r['risk'] == 'BAJO')
medio = sum(1 for r in results if r['risk'] == 'MEDIO')
alto = sum(1 for r in results if r['risk'] == 'ALTO')
print(f"  Thresholds: BAJO<={t1}, MEDIO<={t2}, ALTO>{t2}")
print(f"  BAJO: {bajo} ({bajo/n:.0%}) | MEDIO: {medio} ({medio/n:.0%}) | ALTO: {alto} ({alto/n:.0%})")
for lv in ['BAJO', 'MEDIO', 'ALTO']:
    g = [r for r in results if r['risk'] == lv]
    fs = sum(1 for r in g if r['fpd30'] == 1); fn = sum(1 for r in g if r['fpd30'] == 0); ft = fs + fn
    print(f"  {lv} FPD: {fs}/{ft} = {fs/ft*100:.1f}%" if ft else f"  {lv}: sin FPD")

# EXCEL
wb = Workbook()
hf = Font(bold=True, color="FFFFFF", size=10)
hfl = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
bf = Font(bold=True, size=10)
rf = Font(color="FF0000", bold=True); gf = Font(color="008000", bold=True); yf = Font(color="FF8C00", bold=True)
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def hdr(ws, row, vals):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v); c.font = hf; c.fill = hfl
        c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = tb

# SHEET 1: Detalle
ws = wb.active; ws.title = f"Detalle {n} Casos"
hdrs = ["#", "DNI", "Nombre", "Edad", "Sueldo", "RCI", "Score", "Depto", "Sol ID",
        "Sols", "Firmadas", "Rech", "Docs OCR", "Boletas", "Emp.", "Nm Mis", "DNI Mis", "EQF OCR",
        "DPD", "FPD30+", "Pts", "RIESGO", "Anomalias"]
hdr(ws, 1, hdrs)
results_sorted = sorted(results, key=lambda x: -x['pts'])
for i, r in enumerate(results_sorted, 2):
    vals = [i-1, r['dni'], r['nombre'], r['edad'], r['sueldo'],
            f"{r['rci']:.0%}" if r['rci'] else 'N/A', r['score'] if r['score'] else 'N/A',
            r['depto'], int(r['sol_id']), r['n_sols'], r['n_firmadas'], r['n_rechazadas'],
            r['n_docs_ocr'], r['n_boletas'], r['n_empleadores'], r['n_nombres_mismatch'], r['n_dnis_mismatch'],
            r['equifax_score_ocr'] if r['equifax_score_ocr'] else 'N/A',
            r['dpd2'] if r['dpd2'] is not None else 'N/A',
            'SI' if r['fpd30'] == 1 else ('NO' if r['fpd30'] == 0 else 'N/A'),
            r['pts'], r['risk'], r['anomalies']]
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=i, column=j, value=v); c.border = tb
        if j == 5 and isinstance(v, (int, float)): c.number_format = '#,##0'
        if j == 22:
            if v == 'ALTO': c.font = rf
            elif v == 'MEDIO': c.font = yf
            else: c.font = gf
for c in range(1, len(hdrs)+1): ws.column_dimensions[ws.cell(1, c).column_letter].width = 12
ws.column_dimensions['C'].width = 28
ws.column_dimensions[ws.cell(1, len(hdrs)).column_letter].width = 80

# SHEET 2: Resumen
ws2 = wb.create_sheet("Resumen por Riesgo")
hdr(ws2, 1, ["Riesgo", "Casos", "%", "Medibles *", "Impago **", "Pago OK", "Tasa FPD", "Sueldo Prom", "Score Prom", "Sols Prom"])
for idx, level in enumerate(['BAJO', 'MEDIO', 'ALTO'], 2):
    g = [r for r in results if r['risk'] == level]; gn = len(g)
    fs = sum(1 for r in g if r['fpd30'] == 1); fn = sum(1 for r in g if r['fpd30'] == 0); ft = fs + fn
    sc = [r['score'] for r in g if r['score']]; avg_sc = sum(sc)/len(sc) if sc else 0
    vals = [level, gn, f"{gn/n:.0%}", ft, fs, fn, f"{fs/ft:.1%}" if ft else 'N/A',
            f"S/ {sum(r['sueldo'] for r in g)/gn:,.0f}" if gn else '-',
            f"{avg_sc:.0f}" if avg_sc else 'N/A',
            f"{sum(r['n_sols'] for r in g)/gn:.1f}" if gn else '-']
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=idx, column=j, value=v); c.border = tb
        if j == 1:
            if v == 'ALTO': c.font = rf
            elif v == 'MEDIO': c.font = yf
            else: c.font = gf
ft_all = sum(1 for r in results if r['fpd30'] is not None)
fs_all = sum(1 for r in results if r['fpd30'] == 1); fn_all = sum(1 for r in results if r['fpd30'] == 0)
vals_t = ['TOTAL', n, '100%', fs_all+fn_all, fs_all, fn_all,
          f"{fs_all/(fs_all+fn_all):.1%}" if (fs_all+fn_all) else 'N/A',
          f"S/ {sum(r['sueldo'] for r in results)/n:,.0f}", '', '']
for j, v in enumerate(vals_t, 1): c = ws2.cell(row=5, column=j, value=v); c.border = tb; c.font = bf
for c in range(1, 11): ws2.column_dimensions[ws2.cell(1, c).column_letter].width = 16
ws2.cell(7, 1, value="* Medibles: solicitudes con 2+ meses en cartera").font = Font(italic=True, size=9, color="6B7280")
ws2.cell(8, 1, value="** Impago: FPD 30+ (no pago 1ra cuota, >30 dias atraso)").font = Font(italic=True, size=9, color="6B7280")

# SHEET 3: Anomalias
ws3 = wb.create_sheet("Anomalias Frecuentes")
hdr(ws3, 1, ["Anomalia", "Frecuencia", "%", "En ALTO", "En MEDIO", "Tasa FPD"])
all_a = []
for r in results:
    if r['anomalies'] != 'Ninguna':
        for a in r['anomalies'].split('; '):
            key = re.sub(r'[\d,./]+', 'X', a); all_a.append((key, a))
norm_counts = Counter(k for k, _ in all_a)
for i, (key, cnt) in enumerate(norm_counts.most_common(25), 2):
    example = next(a for k, a in all_a if k == key)
    fg = [r for r in results if any(re.sub(r'[\d,./]+', 'X', a2) == key for a2 in r['anomalies'].split('; '))]
    in_a = sum(1 for r in fg if r['risk'] == 'ALTO'); in_m = sum(1 for r in fg if r['risk'] == 'MEDIO')
    fs = sum(1 for r in fg if r['fpd30'] == 1); fn = sum(1 for r in fg if r['fpd30'] == 0); ft = fs + fn
    for j, v in enumerate([example, cnt, f"{cnt/n:.0%}", in_a, in_m, f"{fs/ft:.1%}" if ft else 'N/A'], 1):
        ws3.cell(row=i, column=j, value=v).border = tb
ws3.column_dimensions['A'].width = 60

# SHEET 4: Casos ALTO
ws4 = wb.create_sheet("Casos Riesgo ALTO")
hdr(ws4, 1, ["#", "DNI", "Nombre", "Edad", "Sueldo", "Score", "Sols", "Rech",
             "OCR", "Bol", "Emp", "NmMis", "DNIMis", "EQF", "DPD", "FPD", "Pts", "Anomalias"])
altos = [r for r in results if r['risk'] == 'ALTO']
altos.sort(key=lambda x: -x['pts'])
for i, r in enumerate(altos, 2):
    vals = [i-1, r['dni'], r['nombre'], r['edad'], r['sueldo'],
            r['score'] if r['score'] else 'N/A', r['n_sols'], r['n_rechazadas'],
            r['n_docs_ocr'], r['n_boletas'], r['n_empleadores'], r['n_nombres_mismatch'], r['n_dnis_mismatch'],
            r['equifax_score_ocr'] if r['equifax_score_ocr'] else 'N/A',
            r['dpd2'] if r['dpd2'] is not None else 'N/A',
            'SI' if r['fpd30'] == 1 else ('NO' if r['fpd30'] == 0 else 'N/A'),
            r['pts'], r['anomalies']]
    for j, v in enumerate(vals, 1):
        c = ws4.cell(row=i, column=j, value=v); c.border = tb
        if j == 5 and isinstance(v, (int, float)): c.number_format = '#,##0'
ws4.column_dimensions['C'].width = 28
ws4.column_dimensions[ws4.cell(1, 18).column_letter].width = 80

wb.save(OUTPUT)
print(f"\nOK: {OUTPUT}")
print(f"Total: {n} | BAJO: {bajo} | MEDIO: {medio} | ALTO: {alto}")
print(f"\nTop 10 ALTO:")
for r in altos[:10]:
    print(f"  Sol {r['sol_id']} | {r['nombre'][:30]:30s} | Pts {r['pts']:2d} | {r['anomalies'][:80]}")
